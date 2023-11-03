from stravalib.client import Client
from stravalib import unithelper
import pickle
import time
from datetime import datetime
import yaml
import openpyxl

# Get config
config = yaml.safe_load(open("../config.yml"))
CLIENT_ID = config["strava_secrets"]["client_id"]
CLIENT_SECRET = config["strava_secrets"]["client_secret"]
CODE = config["strava_secrets"]["code"]

RACE_START = datetime(2023, 8, 24)


def get_milestones():
    # Get Milestones
    spreadsheet = openpyxl.load_workbook("./Running Milestones.xlsx")
    spreadsheet1 = spreadsheet["Sheet1"]

    milestones = {}
    for row in range(2, spreadsheet1.max_row):
        milestones[spreadsheet1.cell(row, 1).value] = spreadsheet1.cell(row, 2).value
    return milestones


def get_nearest_milestones(milestones, total_distance):
    # Get closest and completed milestones
    closest_below = (None, -1)
    closest_above = (None, 50000)
    completed = []

    for milestone, distance in milestones.items():
        if distance < total_distance:
            completed.append((milestone, distance))
            if distance > closest_below[1]:
                closest_below = (milestone, distance)
        elif distance > total_distance:
            if distance < closest_above[1]:
                closest_above = (milestone, distance)

    return closest_below, closest_above, completed


def get_distance_stats(activities):
    total_distance = 0
    run_distance = 0
    cycle_distance = 0
    hike_distance = 0

    activities_since_start = []
    print("Your activities:")
    for activity in activities:
        dt = datetime(activity.start_date.year,
                      activity.start_date.month,
                      activity.start_date.day)
        if dt > RACE_START:
            activities_since_start.append(activity)
            print(activity.name)
            # print(activity.type)
            total_distance += unithelper.kilometer(activity.distance).num
            if activity.type == "Run":
                run_distance += unithelper.kilometer(activity.distance).num
            elif activity.type == "Ride":
                cycle_distance += unithelper.kilometer(activity.distance).num
            elif activity.type == "Hike":
                hike_distance += unithelper.kilometer(activity.distance).num

    return total_distance, run_distance, cycle_distance, hike_distance


# Strava authentication
client = Client()
# url = client.authorization_url(client_id=CLIENT_ID, redirect_uri='http://127.0.0.1:5000/authorization', scope=['read_all','profile:read_all','activity:read_all'])
# print(url)
#
# access_token = client.exchange_code_for_token(client_id=CLIENT_ID, client_secret=CLIENT_SECRET, code=CODE)

# with open('./access_token.pickle', 'wb') as f:
    # pickle.dump(access_token, f)

with open('../access_token.pickle', 'rb') as f:
    access_token = pickle.load(f)

if time.time() > access_token['expires_at']:
    print('Token has expired, will refresh')
    refresh_response = client.refresh_access_token(client_id=CLIENT_ID, client_secret=CLIENT_SECRET,
                                                   refresh_token=access_token["refresh_token"])
    access_token = refresh_response
    with open('../../access_token.pickle', 'wb') as f:
        pickle.dump(refresh_response, f)
    print('Refreshed token saved to file')
    client.access_token = refresh_response['access_token']
    client.refresh_token = refresh_response['refresh_token']
    client.token_expires_at = refresh_response['expires_at']

else:
    print('Token still valid, expires at {}'
          .format(time.strftime("%a, %d %b %Y %H:%M:%S %Z", time.localtime(access_token['expires_at']))))
    client.access_token = access_token['access_token']
    client.refresh_token = access_token['refresh_token']
    client.token_expires_at = access_token['expires_at']

milestones = get_milestones()
print(milestones)

# Get distance data
athlete = client.get_athlete()
print(f"Welcome user {athlete.firstname} {athlete.lastname}")
activities = client.get_activities(limit=100)
total, run, cycle, hike = get_distance_stats(activities)
print(f"You have travelled: {total}km.")
print(f"You have run {run}km, cycled {cycle}km, and hiked {hike}km.")


closest_below, closest_above, completed = get_nearest_milestones(milestones, total)
print(closest_below)
print(closest_above)
print(completed)
