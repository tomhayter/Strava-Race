from django.shortcuts import render
from django.conf import settings
from stravalib.client import Client
from stravalib import unithelper
import pickle
import time as t
from datetime import datetime, timedelta
import yaml
import openpyxl

# Get config
config = yaml.safe_load(open(f"{settings.BASE_DIR}\\..\\Deployment\\config.yml"))

USERS = ["tom", "ben", "justin"]

RACE_START = datetime(2023, 8, 24)

class Statistics:
    def __init__(self, client, activities):
        self.total_distance = 0
        self.run_distance = 0
        self.cycle_distance = 0
        self.hike_distance = 0
        self.total_elevation = 0
        self.total_time = timedelta()
        self.countries = set()
        self.best_5k = timedelta.max
        self.best_10k = timedelta.max
        self.best_half = timedelta.max
        # self.best_marathon = timedelta()

        self.num_activities = 0
        self.longest_run = 0
        self.longest_cycle = 0
        self.longest_hike = 0
        self.largest_climb = 0
        self.highest_point = 0

        for activity in activities:
            dt = datetime(activity.start_date.year,
                        activity.start_date.month,
                        activity.start_date.day)
            if dt < RACE_START:
                continue
            self.total_distance += unithelper.kilometer(activity.distance).num
            self.total_elevation += unithelper.meter(activity.total_elevation_gain).num
            self.largest_climb = max(self.largest_climb, unithelper.meter(activity.total_elevation_gain).num)
            self.total_time += activity.moving_time
            self.num_activities += 1
            self.highest_point = max(self.highest_point, activity.elev_high)
            self.countries.add(activity.location_country)
            if activity.type == "Run":
                self.longest_run = max(self.longest_run, unithelper.kilometer(activity.distance).num)
                self.run_distance += unithelper.kilometer(activity.distance).num
                activity_with_be = client.get_activity(activity.id, True)
                if activity_with_be.best_efforts:
                    for effort in activity_with_be.best_efforts:
                        if effort.name == "5k":
                            self.best_5k = min(self.best_5k, effort.elapsed_time)
                        elif effort.name == "10k":
                            self.best_10k = min(self.best_10k, effort.elapsed_time)
                        elif effort.name == "Half-Marathon":
                            self.best_half = min(self.best_half, effort.elapsed_time)
            elif activity.type == "Ride":
                self.longest_cycle = max(self.longest_cycle, unithelper.kilometer(activity.distance).num)
                self.cycle_distance += unithelper.kilometer(activity.distance).num
            elif activity.type == "Hike" or activity.type == "Walk":
                self.longest_hike = max(self.longest_hike, unithelper.kilometer(activity.distance).num)
                self.hike_distance += unithelper.kilometer(activity.distance).num
                

class User:
    def __init__(self, user):
        self.client = get_client_for_user(user)
        self.athlete = self.client.get_athlete()
        self.activities = list(self.client.get_activities(limit=100, after="2023-08-24T00:00:00Z"))
        self.activities = sorted(self.activities, key=lambda x: x.start_date)
        self.stats = Statistics(self.client, self.activities)


class Cache:
    def __init__(self):
        # TODO: Update cache regularly using last updated time
        self.updated = datetime.now()
        self.users = {}

    def add_user(self, user):
        self.users[user] = User(user)

CACHE = Cache()

class Milestone:
    def __init__(self, name, distance, date_achieved, link):
        self.name = name
        self.distance = distance
        if date_achieved is None:
            self.date_achieved = ""
        else:
            self.date_achieved = datetime.date(date_achieved)
        if link is None:
            self.link = ""
        else: 
            self.link = link

def get_user_column(user):
    # Get the column number for a user
    if user == "tom":
        return 3
    if user == "ben":
        return 4
    if user == "justin":
        return 5

def get_trophies():
    # Get trophies from excel
    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\Deployment\\Running Milestones.xlsx")
    spreadsheet1 = spreadsheet["Trophies"]

    trophies = []
    for row in range(2, spreadsheet1.max_row+1):
        if spreadsheet1.cell(row, 2).value is None:
            continue
        unit = str(spreadsheet1.cell(row, 4).value)
        if unit == "None":
            unit = ""
        holders = spreadsheet1.cell(row, 2).value.split(", ")
        trophies.append({"name": spreadsheet1.cell(row, 1).value,
                        "holders": str([get_athlete(name).firstname for name in holders]).replace("'", "").replace("[", "").replace("]", ""),
                        "value": str(spreadsheet1.cell(row, 3).value) + unit
                        })
        
    return trophies

def get_altitude_milestones(user):
    # Get milestones from excel with following data (milestone name, distance, link, date achieved)
    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\Deployment\\Running Milestones.xlsx")
    spreadsheet1 = spreadsheet["Altitude"]

    user_column = get_user_column(user)

    milestones = []
    for row in range(3, spreadsheet1.max_row+1):
        milestones.append(
            Milestone(
            spreadsheet1.cell(row, 1).value,
            spreadsheet1.cell(row, 2).value,
            spreadsheet1.cell(row, user_column).value,
            spreadsheet1.cell(row, 6).value))
    return milestones

def get_milestones(user):
    # Get milestones from excel with following data (milestone name, distance, link, date achieved)
    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\Deployment\\Running Milestones.xlsx")
    spreadsheet1 = spreadsheet["Distance"]

    user_column = get_user_column(user)

    milestones = []
    for row in range(3, spreadsheet1.max_row+1):
        milestones.append(
            Milestone(
            spreadsheet1.cell(row, 1).value,
            spreadsheet1.cell(row, 2).value,
            spreadsheet1.cell(row, user_column).value,
            spreadsheet1.cell(row, 6).value))
    return milestones

def get_nearest_milestones(milestones, total_distance):
    # Get closest and completed milestones
    closest_below = Milestone(None, -1, None, None)
    closest_above = Milestone(None, 50000, None, None)
    completed = []

    for m in milestones:
        if m.distance < total_distance:
            completed.append(m)
            if m.distance > closest_below.distance:
                closest_below = m
        elif m.distance > total_distance:
            if m.distance < closest_above.distance:
                closest_above = m

    return closest_below, closest_above, completed

def convert_tuple(tuple):
    return (str(tuple[0]).replace("'", "").replace("[", "").replace("]", ""), tuple[1])

def update_trophy_winners():
    most_activities = (None, 0)
    most_running = (None, 0)
    most_cycling = (None, 0)
    most_hiking = (None, 0)
    longest_run = (None, 0)
    longest_cycle = (None, 0)
    longest_hike = (None, 0)
    largest_climb = (None, 0)
    most_climbing = (None, 0)
    most_time = (None, timedelta())
    highest_point = (None, 0)
    most_countries = (None, 0)
    fastest_5k = (None, timedelta.max)
    fastest_10k = (None, timedelta.max)
    fastest_half = (None, timedelta.max)
    for user in USERS:
        stats = get_stats(user)
        if most_activities[1] < stats.num_activities:
            most_activities = ([user], stats.num_activities)
        elif most_activities[1] == stats.num_activities:
            most_activities[0].append(user)

        if most_running[1] < stats.run_distance:
            most_running = ([user], round(stats.run_distance, 2))
        elif most_running[1] == stats.run_distance:
            most_running[0].append(user)

        if most_cycling[1] < stats.cycle_distance:
            most_cycling = ([user], round(stats.cycle_distance, 2))
        elif most_cycling[1] == stats.cycle_distance:
            most_cycling[0].append(user)

        if most_hiking[1] < stats.hike_distance:
            most_hiking = ([user], round(stats.hike_distance, 2))
        elif most_hiking[1] == stats.hike_distance:
            most_hiking[0].append(user)

        if longest_run[1] < stats.longest_run:
            longest_run = ([user], round(stats.longest_run, 2))
        elif longest_run[1] == stats.longest_run:
            longest_run[0].append(user)

        if longest_cycle[1] < stats.longest_cycle:
            longest_cycle = ([user], round(stats.longest_cycle, 2))
        elif longest_cycle[1] == stats.longest_cycle:
            longest_cycle[0].append(user)

        if longest_hike[1] < stats.longest_hike:
            longest_hike = ([user], round(stats.longest_hike, 2))
        elif longest_hike[1] == stats.longest_hike:
            longest_hike[0].append(user)

        if largest_climb[1] < stats.largest_climb:
            largest_climb = ([user], round(stats.largest_climb))
        elif largest_climb[1] == stats.largest_climb:
            largest_climb[0].append(user)

        if most_climbing[1] < stats.total_elevation:
            most_climbing = ([user], round(stats.total_elevation))
        elif most_climbing[1] == stats.total_elevation:
            most_climbing[0].append(user)

        if most_time[1] < stats.total_time:
            most_time = ([user], stats.total_time)
        elif most_time[1] == stats.total_time:
            most_time[0].append(user)

        if highest_point[1] < stats.highest_point:
            highest_point = ([user], round(stats.highest_point))
        elif highest_point[1] == stats.highest_point:
            highest_point[0].append(user)

        if most_countries[1] < len(stats.countries):
            most_countries = ([user], len(stats.countries))
        elif most_countries[1] == len(stats.countries):
            most_countries[0].append(user)

        if fastest_5k[1] > stats.best_5k:
            fastest_5k = ([user], stats.best_5k)
        elif fastest_5k[1] == stats.best_5k:
            fastest_5k[0].append(user)

        if fastest_10k[1] > stats.best_10k:
            fastest_10k = ([user], stats.best_10k)
        elif fastest_10k[1] == stats.best_10k:
            fastest_10k[0].append(user)

        if fastest_half[1] > stats.best_half:
            fastest_half = ([user], stats.best_half)
        elif fastest_half[1] == stats.best_half:
            fastest_half[0].append(user)

    most_activities = convert_tuple(most_activities)
    most_running = convert_tuple(most_running)
    most_cycling = convert_tuple(most_cycling)
    most_hiking = convert_tuple(most_hiking)
    longest_run = convert_tuple(longest_run)
    longest_cycle = convert_tuple(longest_cycle)
    longest_hike = convert_tuple(longest_hike)
    largest_climb = convert_tuple(largest_climb)
    most_climbing = convert_tuple(most_climbing)
    most_time = convert_tuple(most_time)
    highest_point = convert_tuple(highest_point)
    most_countries = convert_tuple(most_countries)
    fastest_5k = convert_tuple(fastest_5k)
    fastest_10k = convert_tuple(fastest_10k)
    fastest_half = convert_tuple(fastest_half)

    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\Deployment\\Running Milestones.xlsx")
    trophy_sheet = spreadsheet["Trophies"]
    trophy_sheet["B2"], trophy_sheet["C2"] = most_climbing
    trophy_sheet["B3"], trophy_sheet["C3"] = longest_run
    trophy_sheet["B4"], trophy_sheet["C4"] = longest_cycle
    trophy_sheet["B5"], trophy_sheet["C5"] = longest_hike
    trophy_sheet["B6"], trophy_sheet["C6"] = largest_climb
    trophy_sheet["B7"], trophy_sheet["C7"] = most_activities
    trophy_sheet["B8"], trophy_sheet["C8"] = most_time
    trophy_sheet["B9"],  trophy_sheet["C9"] = most_running
    trophy_sheet["B10"], trophy_sheet["C10"] = most_cycling
    trophy_sheet["B11"], trophy_sheet["C11"] = most_hiking
    trophy_sheet["B12"], trophy_sheet["C12"] = most_countries
    trophy_sheet["B13"], trophy_sheet["C13"] = fastest_5k
    trophy_sheet["B14"], trophy_sheet["C14"] = fastest_10k
    trophy_sheet["B15"], trophy_sheet["C15"] = fastest_half
    trophy_sheet["B17"], trophy_sheet["C17"] = highest_point

    spreadsheet.save(filename=f"{settings.BASE_DIR}\\..\\Deployment\\Running Milestones.xlsx")
    
def update_milestones():
    # Update the milestone dates achieved
    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\Deployment\\Running Milestones.xlsx")
    spreadsheet1 = spreadsheet["Distance"]
    for user in USERS:
        user_column = get_user_column(user)
        activities = get_activities(user)
        total = get_stats(user).total_distance
        for row in range(3, spreadsheet1.max_row+1):
            milestone_distance = spreadsheet1.cell(row, 2).value
            if milestone_distance > total:
                break
            running_total = 0
            for activity in activities:
                running_total += unithelper.kilometer(activity.distance).num
                if running_total > milestone_distance:
                    spreadsheet1.cell(row, user_column).value = datetime(activity.start_date.year,
                                                                         activity.start_date.month,
                                                                         activity.start_date.day)
                    break
    
    spreadsheet.save(filename=f"{settings.BASE_DIR}\\..\\Deployment\\Running Milestones.xlsx")

def get_client_for_user(user):
    # Get the client for a user.
    CLIENT_ID = config[user]["client_id"]
    CLIENT_SECRET = config[user]["client_secret"]

    # Strava authentication
    client = Client()

    with open(f'{settings.BASE_DIR}\\..\\Deployment\\access_tokens\\{user}_access_token.pickle', 'rb') as f:
        access_token = pickle.load(f)

    if t.time() > access_token['expires_at']:
        print('Token has expired, will refresh')
        refresh_response = client.refresh_access_token(client_id=CLIENT_ID, client_secret=CLIENT_SECRET,
                                                    refresh_token=access_token["refresh_token"])
        access_token = refresh_response
        with open(f'{settings.BASE_DIR}\\..\\Deployment\\access_tokens\\{user}_access_token.pickle', 'wb') as f:
            pickle.dump(refresh_response, f)
        print('Refreshed token saved to file')
        client.access_token = refresh_response['access_token']
        client.refresh_token = refresh_response['refresh_token']
        client.token_expires_at = refresh_response['expires_at']

    else:
        print('Token still valid, expires at {}'
            .format(t.strftime("%a, %d %b %Y %H:%M:%S %Z", t.localtime(access_token['expires_at']))))
        client.access_token = access_token['access_token']
        client.refresh_token = access_token['refresh_token']
        client.token_expires_at = access_token['expires_at']

    return client

def get_stats(user):
    if user not in CACHE.users:
        CACHE.add_user(user)
    return CACHE.users[user].stats

def get_activities(user):
    if user not in CACHE.users:
        CACHE.add_user(user)
    return CACHE.users[user].activities

def get_athlete(user):
    if user not in CACHE.users:
        CACHE.add_user(user)
    return CACHE.users[user].athlete


def home(request):
    rankings = []

    cumulative = 0
    for user in USERS:
        # Get distance data
        athlete = get_athlete(user)
        total = get_stats(user).total_distance
        rankings.append((athlete.firstname, total))
        cumulative += total

    rankings = sorted(rankings, key=lambda x: x[1])

    # TODO: Thread these
    update_milestones()
    update_trophy_winners()

    context = {
        "first": rankings[-1][0],
        "second" : rankings[-2][0],
        "third" : rankings[-3][0],
        "first_distance" : round(rankings[-1][1]),
        "second_distance" : round(rankings[-2][1]),
        "third_distance" : round(rankings[-3][1]),
        "cumulative_distance" : round(cumulative)
    }
    return render(request, "tracker/home.html", context)

def user(request):

    user = request.GET.get("name").lower()
    milestones = get_milestones(user)
    altitude_milestones = get_altitude_milestones(user)

    # Get distance data
    athlete = get_athlete(user)
    stats = get_stats(user)
    total = stats.total_distance
    run = stats.run_distance
    cycle = stats.cycle_distance
    hike = stats.hike_distance
    altitude = stats.total_elevation

    completed_altitudes = [a for a in altitude_milestones if a.distance < altitude]

    trophies = get_trophies()
    my_trophies = []
    for trophy in trophies:
        if user in trophy["holders"].lower():
            my_trophies.append(trophy)
    

    closest_below, closest_above, completed = get_nearest_milestones(milestones, total)

    progress = 100 * (total - closest_below.distance)/(closest_above.distance - closest_below.distance)

    context = {
        "name": athlete.firstname + " " + athlete.lastname,
        "total_distance" : round(total, 2),
        "run_distance" : round(run, 2),
        "cycle_distance" : round(cycle, 2),
        "hike_distance" : round(hike, 2),
        "altitude": round(altitude, 1),
        "last_milestone" : closest_below,
        "next_milestone": closest_above,
        "progress": round(progress),
        "trophies" : my_trophies,
        "completed_milestones": completed,
        "completed_altitude_milestones": completed_altitudes,
        "arg_name": user
    }
    return render(request, "tracker/user.html", context)

def milestones(request):
    user = request.GET.get("name").lower()
    milestones = get_milestones(user)
    context = {
        "milestones": milestones,
        "arg_name": user
    }
    return render(request, "tracker/milestones.html", context)

def trophies(request):
    trophies = get_trophies()
    context = {
        "trophies": trophies
    }
    return render(request, "tracker/trophies.html", context)

def altitude_milestones(request):
    user = request.GET.get("name").lower()
    milestones = get_altitude_milestones(user)
    context = {
        "milestones": milestones,
        "arg_name": user
    }
    return render(request, "tracker/milestones.html", context)