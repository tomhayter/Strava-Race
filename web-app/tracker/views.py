from django.shortcuts import render
from django.conf import settings
from stravalib.client import Client
from stravalib import unithelper
import pickle
import time as t
from datetime import datetime, timedelta
import yaml
import openpyxl

# Get config
config = yaml.safe_load(open(f"{settings.BASE_DIR}\\..\\..\\Deployment\\config.yml"))

USERS = ["tom", "ben", "justin"]

RACE_START = datetime(2023, 8, 24)

class Statistics:
    def __init__(self, activities):
        self.total_distance = 0
        self.run_distance = 0
        self.cycle_distance = 0
        self.hike_distance = 0
        self.total_elevation = 0
        self.total_time = timedelta()

        self.num_activities = 0
        self.num_runs = 0
        self.num_cycles = 0
        self.num_hikes = 0
        self.longest_run = 0
        self.longest_cycle = 0
        self.longest_hike = 0
        self.largest_climb = 0

        for activity in activities:
            dt = datetime(activity.start_date.year,
                        activity.start_date.month,
                        activity.start_date.day)
            if dt > RACE_START:
                self.total_distance += unithelper.kilometer(activity.distance).num
                self.total_elevation += unithelper.meter(activity.total_elevation_gain).num
                self.largest_climb = max(self.largest_climb, unithelper.meter(activity.total_elevation_gain).num)
                self.total_time += activity.moving_time
                self.num_activities += 1
                if activity.type == "Run":
                    self.num_runs += 1
                    self.longest_run = max(self.longest_run, unithelper.kilometer(activity.distance).num)
                    self.run_distance += unithelper.kilometer(activity.distance).num
                elif activity.type == "Ride":
                    self.num_cycles += 1
                    self.longest_cycle = max(self.longest_cycle, unithelper.kilometer(activity.distance).num)
                    self.cycle_distance += unithelper.kilometer(activity.distance).num
                elif activity.type == "Hike" or activity.type == "Walk":
                    self.num_hikes += 1
                    self.longest_hike = max(self.longest_hike, unithelper.kilometer(activity.distance).num)
                    self.hike_distance += unithelper.kilometer(activity.distance).num


class User:
    def __init__(self, user):
        self.client = get_client_for_user(user)
        self.name = self.client.get_athlete()
        self.activities = list(self.client.get_activities(limit=100, after="2023-08-24T00:00:00Z"))
        self.activities = sorted(self.activities, key=lambda x: x.start_date)
        self.stats = Statistics(self.activities)


class Cache:
    def __init__(self):
        self.updated = datetime.now()
        self.users = {}

    def add_user(self, user):
        self.users[user] = User(user)

CACHE = Cache()

class Milestone:
    def __init__(self, name, distance, date_achieved, link):
        self.name = name
        self.distance = distance
        if date_achieved is None:
            self.date_achieved = ""
        else:
            self.date_achieved = datetime.date(date_achieved)
        if link is None:
            self.link = ""
        else: 
            self.link = link

    def dictionary(self):
        return {"name": self.name,
                "distance": self.distance,
                "date_achieved": self.date_achieved,
                "link": self.link}

def get_trophies():
    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\..\\Deployment\\Running Milestones.xlsx")
    spreadsheet1 = spreadsheet["Trophies"]

    trophies = []
    for row in range(2, spreadsheet1.max_row+1):
        if spreadsheet1.cell(row, 2).value is None:
            continue
        unit = str(spreadsheet1.cell(row, 4).value)
        if unit == "None":
            unit = ""
        trophies.append({"name": spreadsheet1.cell(row, 1).value,
                        "holder": get_athlete(spreadsheet1.cell(row, 2).value).firstname,
                        "value": str(spreadsheet1.cell(row, 3).value) + unit
                        })
        
    return trophies

def get_milestones():
    # Get Milestones
    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\..\\Deployment\\Running Milestones.xlsx")
    spreadsheet1 = spreadsheet["Distance"]

    milestones = {}
    for row in range(2, spreadsheet1.max_row+1):
        milestones[spreadsheet1.cell(row, 1).value] = spreadsheet1.cell(row, 2).value
    return milestones

def get_detailed_milestones(user):
    # Get Milestones
    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\..\\Deployment\\Running Milestones.xlsx")
    spreadsheet1 = spreadsheet["Distance"]

    user_column = 3
    if user == "ben":
        user_column = 4
    if user == "justin":
        user_column = 5
        

    milestones = []
    for row in range(3, spreadsheet1.max_row+1):
        milestones.append(
            Milestone(
            spreadsheet1.cell(row, 1).value,
            spreadsheet1.cell(row, 2).value,
            spreadsheet1.cell(row, user_column).value,
            spreadsheet1.cell(row, 6).value))
    return milestones

def get_nearest_milestones(milestones, total_distance):
    # Get closest and completed milestones
    closest_below = (None, -1)
    closest_above = (None, 50000)
    completed = []

    for milestone, distance in milestones.items():
        if distance < total_distance:
            completed.append((milestone, distance))
            if distance > closest_below[1]:
                closest_below = (milestone, distance)
        elif distance > total_distance:
            if distance < closest_above[1]:
                closest_above = (milestone, distance)

    return closest_below, closest_above, completed

def update_trophy_winners():
    most_activities = (None, 0)
    most_runs = (None, 0)
    most_cycles = (None, 0)
    most_hikes = (None, 0)
    longest_run = (None, 0)
    longest_cycle = (None, 0)
    longest_hike = (None, 0)
    largest_climb = (None, 0)
    most_climbing = (None, 0)
    most_time = (None, timedelta())
    for user in USERS:
        stats = get_stats(user)
        if most_activities[1] < stats.num_activities:
            most_activities = (user, stats.num_activities)
        if most_runs[1] < stats.num_runs:
            most_runs = (user, stats.num_runs)
        if most_cycles[1] < stats.num_cycles:
            most_cycles = (user, stats.num_cycles)
        if most_hikes[1] < stats.num_hikes:
            most_hikes = (user, stats.num_hikes)
        if longest_run[1] < stats.longest_run:
            longest_run = (user, round(stats.longest_run, 2))
        if longest_cycle[1] < stats.longest_cycle:
            longest_cycle = (user, round(stats.longest_cycle, 2))
        if longest_hike[1] < stats.longest_hike:
            longest_hike = (user, round(stats.longest_hike, 2))
        if largest_climb[1] < stats.largest_climb:
            largest_climb = (user, round(stats.largest_climb, 1))
        if most_climbing[1] < stats.total_elevation:
            most_climbing = (user, round(stats.total_elevation, 1))
        if most_time[1] < stats.total_time:
            most_time = (user, stats.total_time)

    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\..\\Deployment\\Running Milestones.xlsx")
    trophy_sheet = spreadsheet["Trophies"]
    trophy_sheet["B2"] = most_climbing[0]
    trophy_sheet["C2"] = most_climbing[1]
    trophy_sheet["B3"] = longest_run[0]
    trophy_sheet["C3"] = longest_run[1]
    trophy_sheet["B4"] = longest_cycle[0]
    trophy_sheet["C4"] = longest_cycle[1]
    trophy_sheet["B5"] = longest_hike[0]
    trophy_sheet["C5"] = longest_hike[1]
    trophy_sheet["B6"] = largest_climb[0]
    trophy_sheet["C6"] = largest_climb[1]
    trophy_sheet["B7"] = most_activities[0]
    trophy_sheet["C7"] = most_activities[1]
    trophy_sheet["B8"] = most_time[0]
    trophy_sheet["C8"] = most_time[1]
    trophy_sheet["B9"] = most_runs[0]
    trophy_sheet["C9"] = most_runs[1]
    trophy_sheet["B10"] = most_cycles[0]
    trophy_sheet["C10"] = most_cycles[1]
    trophy_sheet["B11"] = most_hikes[0]
    trophy_sheet["C11"] = most_hikes[1]

    spreadsheet.save(filename=f"{settings.BASE_DIR}\\..\\..\\Deployment\\Running Milestones.xlsx")
    
def update_milestones():
    spreadsheet = openpyxl.load_workbook(f"{settings.BASE_DIR}\\..\\..\\Deployment\\Running Milestones.xlsx")
    spreadsheet1 = spreadsheet["Distance"]
    for user in USERS:
        user_column = 3
        if user == "ben":
            user_column = 4
        if user == "justin":
            user_column = 5
        activities = get_activities(user)
        total = get_stats(user).total_distance
        for row in range(3, spreadsheet1.max_row+1):
            distance = spreadsheet1.cell(row, 2).value
            if distance > total:
                break
            running_total = 0
            for activity in activities:
                running_total += unithelper.kilometer(activity.distance).num
                if running_total > distance:
                    spreadsheet1.cell(row, user_column).value = datetime(activity.start_date.year,
                                                                         activity.start_date.month,
                                                                         activity.start_date.day)
                    break
    
    spreadsheet.save(filename=f"{settings.BASE_DIR}\\..\\..\\Deployment\\Running Milestones.xlsx")

def get_client_for_user(user):
    CLIENT_ID = config[user]["client_id"]
    CLIENT_SECRET = config[user]["client_secret"]
    CODE = config[user]["code"]

    # Strava authentication
    client = Client()

    with open(f'{settings.BASE_DIR}\\..\\..\\Deployment\\access_tokens\\{user}_access_token.pickle', 'rb') as f:
        access_token = pickle.load(f)

    if t.time() > access_token['expires_at']:
        print('Token has expired, will refresh')
        refresh_response = client.refresh_access_token(client_id=CLIENT_ID, client_secret=CLIENT_SECRET,
                                                    refresh_token=access_token["refresh_token"])
        access_token = refresh_response
        with open(f'{settings.BASE_DIR}\\..\\..\\Deployment\\access_tokens\\{user}_access_token.pickle', 'wb') as f:
            pickle.dump(refresh_response, f)
        print('Refreshed token saved to file')
        client.access_token = refresh_response['access_token']
        client.refresh_token = refresh_response['refresh_token']
        client.token_expires_at = refresh_response['expires_at']

    else:
        print('Token still valid, expires at {}'
            .format(t.strftime("%a, %d %b %Y %H:%M:%S %Z", t.localtime(access_token['expires_at']))))
        client.access_token = access_token['access_token']
        client.refresh_token = access_token['refresh_token']
        client.token_expires_at = access_token['expires_at']

    return client

def get_stats(user):
    if user not in CACHE.users:
        print("called cache")
        CACHE.add_user(user)
    return CACHE.users[user].stats

def get_activities(user):
    if user not in CACHE.users:
        print("called cache")
        CACHE.add_user(user)
    return CACHE.users[user].activities

def get_athlete(user):
    if user not in CACHE.users:
        print("called cache")
        CACHE.add_user(user)
    return CACHE.users[user].name


def home(request):
    rankings = []

    cumulative = 0
    for user in USERS:
        # Get distance data
        athlete = get_athlete(user)
        total = get_stats(user).total_distance
        rankings.append((athlete.firstname, total))
        cumulative += total

    rankings = sorted(rankings, key=lambda x: x[1])

    update_milestones()
    update_trophy_winners()

    context = {
        "first": rankings[-1][0],
        "second" : rankings[-2][0],
        "third" : rankings[-3][0],
        "first_distance" : round(rankings[-1][1]),
        "second_distance" : round(rankings[-2][1]),
        "third_distance" : round(rankings[-3][1]),
        "cumulative_distance" : round(cumulative)
    }
    return render(request, "tracker/home.html", context)

def user(request):

    user = request.GET.get("name").lower()
    milestones = get_milestones()

    # Get distance data
    athlete = get_athlete(user)
    stats = get_stats(user)
    total = stats.total_distance
    run = stats.run_distance
    cycle = stats.cycle_distance
    hike = stats.hike_distance
    altitude = stats.total_elevation

    trophies = get_trophies()
    my_trophies = []
    for trophy in trophies:
        if trophy["holder"].lower() == user:
            my_trophies.append(trophy)
    

    closest_below, closest_above, completed = get_nearest_milestones(milestones, total)
    detailed_milestones = get_detailed_milestones(user)
    completed2 = [m[0] for m in completed]
    my_milestones = []
    for m in detailed_milestones:
        if m.name in completed2:
            my_milestones.append(m)


    progress = 100 * (total - closest_below[1])/(closest_above[1] - closest_below[1])

    context = {
        "name": athlete.firstname + " " + athlete.lastname,
        "total_distance" : round(total, 2),
        "run_distance" : round(run, 2),
        "cycle_distance" : round(cycle, 2),
        "hike_distance" : round(hike, 2),
        "altitude": round(altitude, 1),
        "last_milestone_name" : closest_below[0],
        "last_milestone_distance": closest_below[1],
        "next_milestone_name": closest_above[0],
        "next_milestone_distance": closest_above[1],
        "progress": round(progress),
        "trophies" : my_trophies,
        "completed_milestones": my_milestones,
        "arg_name": user
    }
    return render(request, "tracker/user.html", context)

def milestones(request):
    user = request.GET.get("name").lower()
    milestones = get_detailed_milestones(user)
    context = {
        "milestones": [i.dictionary() for i in milestones],
        "arg_name": user
    }
    return render(request, "tracker/milestones.html", context)

def trophies(request):
    trophies = get_trophies()
    context = {
        "trophies": trophies
    }
    return render(request, "tracker/trophies.html", context)